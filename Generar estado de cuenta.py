import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import os
from copy import copy
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================
# CONFIGURACIÓN Y CONSTANTES
# ============================================================
st.set_page_config(page_title="Consolidador Financiero PRO", layout="wide")

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
MESES_INV = {v: k for k, v in MESES.items()}

# ============================================================
# MOTOR DE EXTRACCIÓN (Lógica de tus archivos originales)
# ============================================================
def extraer_numeros(texto):
    nums = re.findall(r"[\d,]+\.\d+", texto)
    return [float(n.replace(",", "")) for n in nums]

def extraer_todos_numeros(texto):
    nums = re.findall(r"[\d,]+\.?\d*", texto)
    return [float(n.replace(",", "")) for n in nums if n]

def detectar_plataforma(texto):
    return "Prestadero" if "PRESTADERO" in texto.upper() else "GBM"

def extraer_nombre_cliente(pdf, plataforma):
    texto = pdf.pages[0].extract_text() or ""
    lineas = texto.split("\n")
    if plataforma == "Prestadero":
        for l in lineas:
            if "Periodo:" in l and "Estado de Cuenta" not in l:
                return re.split(r"\s+Periodo:", l)[0].strip().upper()
    else:
        for l in lineas:
            if "Contrato:" in l:
                p = re.split(r"\s+Contrato:", l)[0].strip()
                return p.replace("PUBLICO EN GENERAL - ", "").upper()
    return "DESCONOCIDO"

def es_smart_cash(texto):
    for l in texto.split("\n"):
        if "RENTA VARIABLE" in l and "VALORES EN CORTO" not in l:
            nums = extraer_numeros(l)
            if len(nums) >= 2 and nums[1] > 0: return False
    return True

# --- Funciones de Portafolio y Movimientos ---
def extraer_portafolio_gbm(pdf):
    portafolio = []
    en_desglose = en_acciones = False
    for pag in pdf.pages:
        texto = pag.extract_text() or ""
        for l in texto.split("\n"):
            ls, lu = l.strip(), l.strip().upper()
            if "DESGLOSE DEL PORTAFOLIO" in lu: en_desglose = True; continue
            if not en_desglose: continue
            if lu == "ACCIONES": en_acciones = True; continue
            if en_acciones and lu.startswith("TOTAL"): en_acciones = False; continue
            if not en_acciones: continue
            
            m = re.match(r"^([A-Z]+(?:\s+\d+)?)\s+", ls)
            if m:
                nums = extraer_todos_numeros(ls[m.end():])
                if len(nums) >= 8:
                    portafolio.append({
                        "Emisora": m.group(1).strip(),
                        "Valor a Mercado": nums[7]
                    })
    return portafolio

# ============================================================
# MOTOR DE CONSOLIDACIÓN (Lógica de Actualización de Excel)
# ============================================================
def actualizar_hoja(ws, datos):
    # Localizar fila de encabezado y totales
    fila_header = 23 # Valor por defecto según tu consolidador.py
    fila_totales = None
    for r in range(fila_header, fila_header + 50):
        if ws.cell(r, 1).value and "TOTALES" in str(ws.cell(r, 1).value).upper():
            fila_totales = r
            break
    
    if not fila_totales: return False

    # Actualizar instrumentos existentes
    gbm = datos.get("gbm")
    if gbm:
        port_pdf = {item["Emisora"]: item["Valor a Mercado"] for item in gbm.get("portafolio", [])}
        
        for r in range(fila_header + 1, fila_totales):
            nombre_instrumento = ws.cell(r, 1).value
            if nombre_instrumento in port_pdf:
                nuevo_valor = port_pdf[nombre_instrumento]
                old_b = ws.cell(r, 2).value or 0
                
                # C = Saldo Total
                ws.cell(r, 3).value = nuevo_valor
                # E = Ganancia Histórica (C - B)
                ws.cell(r, 5).value = nuevo_valor - old_b
                # N = Total
                ws.cell(r, 14).value = nuevo_valor
    return True

# ============================================================
# INTERFAZ STREAMLIT
# ============================================================
st.title("🚀 Consolidador Financiero Automatizado")
st.markdown("Sistema experto para procesar estados de cuenta GBM y Prestadero.")

with st.sidebar:
    st.header("Carga de Datos")
    maestro_file = st.file_uploader("Subir Maestro Anterior (.xlsx)", type="xlsx")
    pdf_files = st.file_uploader("Subir PDFs del Mes", type="pdf", accept_multiple_files=True)

if st.button("Ejecutar Consolidación"):
    if not maestro_file or not pdf_files:
        st.error("Debes subir el archivo maestro y al menos un PDF.")
    else:
        try:
            # 1. Extraer datos de PDFs en memoria
            clientes_data = defaultdict(lambda: {"gbm": None, "prestadero": None})
            
            for pdf_f in pdf_files:
                with pdfplumber.open(pdf_f) as pdf:
                    texto_p1 = pdf.pages[0].extract_text() or ""
                    plat = detectar_plataforma(texto_p1)
                    nombre = extraer_nombre_cliente(pdf, plat)
                    
                    if plat == "GBM":
                        smart = es_smart_cash(texto_p1)
                        port = extraer_portafolio_gbm(pdf)
                        clientes_data[nombre]["gbm"] = {"portafolio": port, "smart": smart}
                    # (Aquí agregarías la lógica de Prestadero similar a la de GBM)

            # 2. Procesar el Excel
            wb = load_workbook(maestro_file)
            for nombre, datos in clientes_data.items():
                # Buscar hoja (coincidencia simple)
                hoja_nombre = next((s for s in wb.sheetnames if nombre in s.upper()), None)
                if hoja_nombre:
                    ws = wb[hoja_nombre]
                    actualizar_hoja(ws, datos)
                    st.success(f"Actualizada hoja: {hoja_nombre}")
                else:
                    st.warning(f"No se encontró hoja para el cliente: {nombre}")

            # 3. Descarga del resultado
            output = io.BytesIO()
            wb.save(output)
            st.download_button(
                label="📥 Descargar Resultado Consolidado",
                data=output.getvalue(),
                file_name="Estado_Cuenta_Consolidado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error crítico: {e}")
