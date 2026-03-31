import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
from copy import copy
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ============================================================
# 1. CONFIGURACIÓN Y ESCUDOS CONTRA ERRORES
# ============================================================
st.set_page_config(page_title="Consolidador Confidelis", layout="wide")

def normalizar(t):
    return str(t).upper().strip() if t else ""

def limpiar_numero(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace(",", "").replace("$", "").strip()
    if s in ("", "-", "–", "NA", "N/A", "ND"): return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

# --- ESCUDO DEFINITIVO CONTRA EL BUG DE CELDAS COMBINADAS ---
def leer_celda_segura(ws, row, col):
    """Lee el valor evadiendo el error de celda combinada."""
    celda = ws.cell(row, col)
    if type(celda).__name__ == 'MergedCell':
        for rng in ws.merged_cells.ranges:
            if rng.min_col <= col <= rng.max_col and rng.min_row <= row <= rng.max_row:
                return ws.cell(rng.min_row, rng.min_col).value
    return celda.value

def escribir_celda_segura(ws, row, col, valor):
    """Escribe forzosamente: Descombina, Escribe y Recombina para vencer a openpyxl."""
    celda = ws.cell(row, col)
    if type(celda).__name__ == 'MergedCell':
        for rng in list(ws.merged_cells.ranges):
            if rng.min_col <= col <= rng.max_col and rng.min_row <= row <= rng.max_row:
                rango_str = str(rng)
                try:
                    ws.unmerge_cells(rango_str)
                    ws.cell(rng.min_row, rng.min_col).value = valor
                    ws.merge_cells(rango_str)
                except Exception:
                    pass 
                return
    else:
        celda.value = valor

def clonar_formato(ws, fila_origen, fila_destino):
    """Clona el ADN de la fila (bordes, fuentes) para que no queden filas en blanco."""
    for col in range(1, 16): 
        try:
            c_origen = ws.cell(fila_origen, col)
            c_destino = ws.cell(fila_destino, col)
            if c_origen.has_style:
                c_destino.font = copy(c_origen.font)
                c_destino.border = copy(c_origen.border)
                c_destino.fill = copy(c_origen.fill)
                c_destino.number_format = c_origen.number_format
                c_destino.alignment = copy(c_origen.alignment)
        except AttributeError:
            pass

# ============================================================
# 2. MOTOR DE EXTRACCIÓN (PDF -> DATOS)
# ============================================================
def extraer_numeros(texto):
    nums = re.findall(r"\d[\d,]*\.\d+", texto)
    return [float(n.replace(",", "")) for n in nums if n.replace(",", "")]

def extraer_todos_numeros(texto):
    nums = re.findall(r"\d[\d,]*\.?\d*", texto)
    return [float(n.replace(",", "")) for n in nums if n.replace(",", "")]

def extraer_numero_despues_de(texto, clave):
    idx = texto.find(clave)
    if idx == -1: return 0.0
    sub = texto[idx + len(clave):]
    nums = extraer_numeros(sub)
    return nums[0] if nums else 0.0

def extraer_periodo(texto):
    m = re.search(r"DE\s+([A-Z]+)\s+DE\s+(\d{4})", texto.upper())
    return f"{m.group(1)} {m.group(2)}" if m else "NUEVO"

def procesar_pdf_financiero(f):
    with pdfplumber.open(f) as pdf:
        texto_p1 = pdf.pages[0].extract_text() or ""
        plataforma = "Prestadero" if "PRESTADERO" in texto_p1.upper() else "GBM"
        nombre = "DESCONOCIDO"
        periodo_str = extraer_periodo(texto_p1)
        datos = {}

        if plataforma == "Prestadero":
            for l in texto_p1.split("\n"):
                if "Periodo:" in l and "Estado de Cuenta" not in l:
                    nombre = re.split(r"\s+Periodo:", l)[0].strip().upper()
            datos = {
                "plataforma": "Prestadero",
                "valor_total": extraer_numero_despues_de(texto_p1, "Valor de la Cuenta:"),
                "depositos": extraer_numero_despues_de(texto_p1, "Abonos:"),
                "retiros": extraer_numero_despues_de(texto_p1, "Retiros:"),
                "interes_mes": 0.0
            }
            for l in texto_p1.split("\n"):
                if "Interés Recibido" in l or "Interes Recibido" in l:
                    nums = extraer_numeros(l)
                    if nums: datos["interes_mes"] = nums[0]
        else:
            for l in texto_p1.split("\n"):
                if "Contrato:" in l:
                    nombre = re.split(r"\s+Contrato:", l)[0].replace("PUBLICO EN GENERAL - ", "").strip().upper()
            
            portafolio = []
            movs = {"COMPRAS": defaultdict(float), "VENTAS": defaultdict(float)}
            efectivo_total = 0.0
            en_acciones = en_mov = False
            
            for pag in pdf.pages:
                texto_pag = pag.extract_text() or ""
                if "VALOR TOTAL" in texto_pag.upper() or "EFECTIVO" in texto_pag.upper():
                    nums = extraer_todos_numeros(texto_pag)
                    if len(nums) > 0: efectivo_total = nums[-1]
                
                for ls in texto_pag.split("\n"):
                    lu = ls.strip().upper()
                    if "ACCIONES" in lu: en_acciones = True; continue
                    if en_acciones and (lu.startswith("TOTAL") or "RENDIMIENTO" in lu): en_acciones = False; continue
                    if en_acciones:
                        m = re.match(r"^([A-Z]+(?:\s+\d+)?)\s+", ls.strip())
                        if m:
                            nums = extraer_todos_numeros(ls[m.end():])
                            if len(nums) >= 8:
                                portafolio.append({"emisora": m.group(1).strip(), "valor_mercado": nums[7]})
                    
                    if "DESGLOSE DE MOVIMIENTOS" in lu: en_mov = True; continue
                    if en_mov and "RENDIMIENTO" in lu: en_mov = False; continue
                    if en_mov:
                        if "COMPRA DE ACCIONES" in lu:
                            m = re.search(r"ACCIONES\.\s+([A-Z0-9\s]+?)\s+[\d,]", ls, re.I)
                            nums = extraer_todos_numeros(ls)
                            if m and len(nums) >= 6: movs["COMPRAS"][m.group(1).strip().upper()] += nums[-2]
                        elif "VENTA DE ACCIONES" in lu:
                            m = re.search(r"ACCIONES\.\s+([A-Z0-9\s]+?)\s+[\d,]", ls, re.I)
                            nums = extraer_todos_numeros(ls)
                            if m and len(nums) >= 6: movs["VENTAS"][m.group(1).strip().upper()] += nums[-2]

            datos = {
                "plataforma": "GBM", "periodo": periodo_str,
                "portafolio": portafolio, "movs": movs, "efectivo_total": efectivo_total
            }
        return nombre, datos

# ============================================================
# 3. LÓGICA MAESTRA DE CONSOLIDACIÓN (1 SOLO ARCHIVO)
# ============================================================
def actualizar_hoja_maestra(ws, info):
    fila_header = 23
    fila_totales = fila_efectivo_gbm = None
    
    for r in range(1, 150):
        val = normalizar(leer_celda_segura(ws, r, 1))
        if "INSTRUMENTO" in val: fila_header = r
        if "EFECTIVO GBM" in val: fila_efectivo_gbm = r
        if "TOTALES" in val: fila_totales = r; break
    
    if not fila_totales: return

    # --------------------------------------------------------
    # CASO A: PRESTADERO
    # --------------------------------------------------------
    if info["plataforma"] == "Prestadero":
        for r in range(fila_header + 1, fila_totales):
            if "PRESTADERO" in normalizar(leer_celda_segura(ws, r, 1)):
                saldo_ini = limpiar_numero(leer_celda_segura(ws, r, 2))
                escribir_celda_segura(ws, r, 3, info["valor_total"])
                escribir_celda_segura(ws, r, 5, info["valor_total"] - saldo_ini)
                escribir_celda_segura(ws, r, 7, info["interes_mes"])
                escribir_celda_segura(ws, r, 9, "ESPECULATIVO\nMODERADO")
                escribir_celda_segura(ws, r, 10, info["retiros"])
                escribir_celda_segura(ws, r, 11, info["depositos"])
                escribir_celda_segura(ws, r, 13, "BAJA")
                escribir_celda_segura(ws, r, 14, info["valor_total"])

    # --------------------------------------------------------
    # CASO B: GBM
    # --------------------------------------------------------
    elif info["plataforma"] == "GBM":
        port_pdf = info["portafolio"]
        movs = info["movs"]
        emisoras_en_pdf = {normalizar(i["emisora"]): i["valor_mercado"] for i in port_pdf}
        instrumentos_vistos = set()
        
        total_compras_mes = sum(movs["COMPRAS"].values())
        total_ventas_mes = sum(movs["VENTAS"].values())

        # 1. ELIMINAR FIBRAS ZOMBIES (De abajo hacia arriba para no dañar los índices)
        for r in range(fila_totales - 1, fila_header, -1):
            nom = normalizar(leer_celda_segura(ws, r, 1))
            if not nom or nom == "-" or "EFECTIVO GBM" in nom: continue
            
            old_b = limpiar_numero(leer_celda_segura(ws, r, 2))
            compra = movs["COMPRAS"].get(nom, 0.0)
            venta = movs["VENTAS"].get(nom, 0.0)
            nuevo_c = emisoras_en_pdf.get(nom, 0.0)
            
            if nuevo_c == 0.0 and compra == 0.0 and venta == 0.0 and old_b == 0.0:
                ws.delete_rows(r)
                fila_totales -= 1
                if fila_efectivo_gbm and r < fila_efectivo_gbm:
                    fila_efectivo_gbm -= 1

        # 2. ACTUALIZAR EXISTENTES
        for r in range(fila_header + 1, fila_totales):
            nom = normalizar(leer_celda_segura(ws, r, 1))
            if not nom or nom == "-" or "EFECTIVO GBM" in nom: continue
            
            instrumentos_vistos.add(nom)
            
            old_b = limpiar_numero(leer_celda_segura(ws, r, 2))
            old_c = limpiar_numero(leer_celda_segura(ws, r, 3)) # Memoria del mes anterior
            
            compra = movs["COMPRAS"].get(nom, 0.0)
            venta = movs["VENTAS"].get(nom, 0.0)
            nuevo_c = emisoras_en_pdf.get(nom, 0.0)

            if nom not in emisoras_en_pdf and venta == 0 and compra == 0: continue

            nuevo_b = max(0.0, old_b + compra - venta)
            
            escribir_celda_segura(ws, r, 2, nuevo_b)
            escribir_celda_segura(ws, r, 3, nuevo_c)
            escribir_celda_segura(ws, r, 5, nuevo_c - nuevo_b)
            escribir_celda_segura(ws, r, 7, nuevo_c - old_c + venta - compra) # Cálculo Real del Mes
            escribir_celda_segura(ws, r, 10, venta)
            escribir_celda_segura(ws, r, 11, compra)
            escribir_celda_segura(ws, r, 14, nuevo_c)

        # 3. INSERTAR FIBRAS NUEVAS (Con Formato)
        for item in port_pdf:
            emisora_pdf = normalizar(item["emisora"])
            if emisora_pdf not in instrumentos_vistos:
                pos = fila_efectivo_gbm if fila_efectivo_gbm else fila_totales
                ws.insert_rows(pos)
                
                clonar_formato(ws, pos - 1, pos)
                
                compra_nueva = movs["COMPRAS"].get(emisora_pdf, item["valor_mercado"])
                
                escribir_celda_segura(ws, pos, 1, emisora_pdf)
                escribir_celda_segura(ws, pos, 2, compra_nueva)
                escribir_celda_segura(ws, pos, 3, item["valor_mercado"])
                escribir_celda_segura(ws, pos, 4, info["periodo"])
                escribir_celda_segura(ws, pos, 5, item["valor_mercado"] - compra_nueva)
                escribir_celda_segura(ws, pos, 7, item["valor_mercado"] - compra_nueva)
                escribir_celda_segura(ws, pos, 9, "ESPECULATIVO\nCONSERVADOR")
                escribir_celda_segura(ws, pos, 10, 0.0)
                escribir_celda_segura(ws, pos, 11, compra_nueva)
                escribir_celda_segura(ws, pos, 13, "ALTA")
                escribir_celda_segura(ws, pos, 14, item["valor_mercado"])
                escribir_celda_segura(ws, pos, 15, "GBM")
                
                if fila_efectivo_gbm: fila_efectivo_gbm += 1
                fila_totales += 1

        # 4. BALANCEAR EFECTIVO GBM
        if fila_efectivo_gbm:
            old_efec_b = limpiar_numero(leer_celda_segura(ws, fila_efectivo_gbm, 2))
            
            escribir_celda_segura(ws, fila_efectivo_gbm, 2, old_efec_b + total_ventas_mes - total_compras_mes)
            escribir_celda_segura(ws, fila_efectivo_gbm, 3, info["efectivo_total"])
            escribir_celda_segura(ws, fila_efectivo_gbm, 5, "-")
            escribir_celda_segura(ws, fila_efectivo_gbm, 7, "-")
            escribir_celda_segura(ws, fila_efectivo_gbm, 9, "CONSERVADOR\nCONSERVADOR")
            escribir_celda_segura(ws, fila_efectivo_gbm, 10, total_compras_mes)
            escribir_celda_segura(ws, fila_efectivo_gbm, 11, total_ventas_mes)
            escribir_celda_segura(ws, fila_efectivo_gbm, 13, "ALTA")
            escribir_celda_segura(ws, fila_efectivo_gbm, 14, info["efectivo_total"])

        # 5. ACTUALIZAR FÓRMULAS
        for r in range(fila_header + 1, fila_totales):
            celda_nombre = normalizar(leer_celda_segura(ws, r, 1))
            if celda_nombre and celda_nombre != "-":
                escribir_celda_segura(ws, r, 12, f"=C{r}/C{fila_totales}")
        
        rango_suma = f"{fila_header + 1}:{fila_totales - 1}"
        escribir_celda_segura(ws, fila_totales, 2, f"=SUM(B{rango_suma})")
        escribir_celda_segura(ws, fila_totales, 3, f"=SUM(C{rango_suma})")
        escribir_celda_segura(ws, fila_totales, 5, f"=SUM(E{rango_suma})")
        escribir_celda_segura(ws, fila_totales, 7, f"=SUM(G{rango_suma})")

# ============================================================
# 4. INTERFAZ STREAMLIT
# ============================================================
def main():
    st.title("🏦 Consolidación Financiera a un Clic")
    st.markdown("Carga el Maestro del mes pasado y los PDFs nuevos. El sistema calculará los rendimientos automáticamente.")
    
    col1, col2 = st.columns(2)
    with col1: 
        maestro_file = st.file_uploader("1. Excel Maestro (Mes Anterior)", type="xlsx")
    with col2: 
        pdf_files = st.file_uploader("2. PDFs del Mes Actual", type="pdf", accept_multiple_files=True)

    if st.button("🚀 Iniciar Consolidación", use_container_width=True):
        if maestro_file and pdf_files:
            try:
                wb = load_workbook(maestro_file)
                
                with st.spinner("Sincronizando movimientos y calculando rendimientos..."):
                    for f in pdf_files:
                        nombre, datos = procesar_pdf_financiero(f)
                        sheet_name = next((s for s in wb.sheetnames if all(p in s.upper() for p in nombre.split()[:2])), None)
                        
                        if sheet_name:
                            actualizar_hoja_maestra(wb[sheet_name], datos)
                            st.success(f"✅ Hoja consolidada: {sheet_name}")
                        else:
                            st.warning(f"⚠️ Cliente no encontrado en el Maestro: {nombre}")

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                st.download_button(
                    label="📥 Descargar Estado de Cuenta Consolidado",
                    data=buf,
                    file_name="Estado_Cuenta_Final_Actualizado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                import traceback
                error_detallado = traceback.format_exc()
                st.error(f"❌ Error crítico procesando los datos: {e}")
                with st.expander("Ver detalles técnicos del error"):
                    st.text(error_detallado)
        else:
            st.error("Por favor, sube el Excel y al menos un PDF.")

if __name__ == "__main__":
    main()
